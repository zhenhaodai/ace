#!/usr/bin/env python3
"""
评测脚本：比较生成的答案与标准答案
支持多种数据集格式和评测指标
"""

import argparse
import json
import re
import string
from collections import Counter
from typing import List, Dict, Any, Tuple, Optional
from utils import load_jsonl

# Excel 相关导入（可选）
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def normalize_verdict(s: str) -> str:
    """
    标准化事实核查的判定结果（支持中英文）

    映射规则：
    - T/True/成立/正确/支持 → T
    - F/False/不成立/错误/不支持 → F
    - uncertain/不确定/证据不足/无法判断 → uncertain
    """
    s = s.strip().lower()

    # True/成立 相关
    if s in ['t', 'true', '成立', '正确', '支持', 'yes']:
        return 'T'

    # False/不成立 相关
    if s in ['f', 'false', '不成立', '错误', '不支持', 'no']:
        return 'F'

    # Uncertain/不确定 相关
    if s in ['uncertain', 'u', '不确定', '证据不足', '无法判断', '未知']:
        return 'uncertain'

    # 如果包含关键词（注意：要先检查否定形式，避免误匹配）
    s_clean = s.replace(' ', '').replace('\n', '').replace('\t', '')

    # 先检查否定形式和不确定形式
    if any(word in s_clean for word in ['不成立', '不支持', '不正确', '错误', 'false', 'notrue']):
        return 'F'
    if any(word in s_clean for word in ['不确定', '证据不足', '无法判断', 'uncertain']):
        return 'uncertain'
    # 再检查肯定形式
    if any(word in s_clean for word in ['成立', '正确', '支持', 'true']):
        return 'T'

    # 如果无法识别，返回原始小写文本
    return s


def normalize_answer(s: str) -> str:
    """标准化答案：小写化、去标点、去冠词、去多余空格"""
    def remove_articles(text):
        return re.sub(r'\b(a|an|the)\b', ' ', text)

    def white_space_fix(text):
        return ' '.join(text.split())

    def remove_punc(text):
        exclude = set(string.punctuation)
        return ''.join(ch for ch in text if ch not in exclude)

    def lower(text):
        return text.lower()

    return white_space_fix(remove_articles(remove_punc(lower(s))))


def extract_answer_from_tags(text: str) -> str:
    """从 <answer></answer> 标签中提取答案"""
    match = re.search(r'<answer>\s*(.*?)\s*</answer>', text, re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()
    # 如果没有标签，返回整个文本
    return text.strip()


def is_fact_checking_dataset(dataset: str) -> bool:
    """判断是否是事实核查数据集"""
    fact_checking_datasets = ["hover", "exfever", "fever", "factcheck"]
    return dataset.lower() in fact_checking_datasets or "fact" in dataset.lower()


def compute_exact_match(prediction: str, ground_truth: str, is_fact_check: bool = False) -> float:
    """计算精确匹配（Exact Match）"""
    if is_fact_check:
        # 对于事实核查任务，使用判定结果标准化
        return float(normalize_verdict(prediction) == normalize_verdict(ground_truth))
    else:
        # 对于问答任务，使用传统标准化
        return float(normalize_answer(prediction) == normalize_answer(ground_truth))


def compute_f1(prediction: str, ground_truth: str, is_fact_check: bool = False) -> float:
    """计算 F1 分数"""
    if is_fact_check:
        # 对于事实核查任务，F1 等同于 EM（因为是分类任务）
        return compute_exact_match(prediction, ground_truth, is_fact_check=True)

    pred_tokens = normalize_answer(prediction).split()
    truth_tokens = normalize_answer(ground_truth).split()

    # 如果都为空，返回 1.0
    if len(pred_tokens) == 0 and len(truth_tokens) == 0:
        return 1.0

    # 如果其中一个为空，返回 0.0
    if len(pred_tokens) == 0 or len(truth_tokens) == 0:
        return 0.0

    common_tokens = Counter(pred_tokens) & Counter(truth_tokens)
    num_common = sum(common_tokens.values())

    if num_common == 0:
        return 0.0

    precision = num_common / len(pred_tokens)
    recall = num_common / len(truth_tokens)
    f1 = 2 * precision * recall / (precision + recall)

    return f1


def compute_yes_no_accuracy(prediction: str, ground_truth: str) -> float:
    """计算 Yes/No 准确率（用于 hover, exfever 等数据集）"""
    pred_normalized = normalize_answer(prediction)
    truth_normalized = normalize_answer(ground_truth)

    # 提取 Yes/No
    pred_answer = None
    if 'yes' in pred_normalized:
        pred_answer = 'yes'
    elif 'no' in pred_normalized:
        pred_answer = 'no'

    truth_answer = None
    if 'yes' in truth_normalized:
        truth_answer = 'yes'
    elif 'no' in truth_normalized:
        truth_answer = 'no'

    if pred_answer is None or truth_answer is None:
        return 0.0

    return float(pred_answer == truth_answer)


def get_ground_truth_answer(item: Dict[str, Any], dataset: str) -> str:
    """根据数据集类型提取标准答案"""
    # 中文数据集字段（优先检查）
    # 检查 original_row 中的中文字段
    if "original_row" in item:
        original = item["original_row"]
        # 尝试各种可能的中文字段名
        for field in ["人工评测结果", "标准答案", "答案", "label"]:
            if field in original and original[field]:
                answer = original[field]
                # 跳过 NaN 值
                if answer != answer:  # NaN check
                    continue
                return str(answer)

    # 直接检查根节点的中文字段
    for field in ["人工评测结果", "标准答案", "答案"]:
        if field in item and item[field]:
            answer = item[field]
            if answer != answer:  # NaN check
                continue
            return str(answer)

    # 尝试多种可能的英文字段名
    if "answer" in item:
        return str(item["answer"])

    if "answers" in item:
        answers = item["answers"]
        if isinstance(answers, list) and len(answers) > 0:
            return str(answers[0])
        return str(answers)

    if "answers_objects" in item:
        # Bamboogle 等数据集格式
        answers_obj = item["answers_objects"]
        if isinstance(answers_obj, list) and len(answers_obj) > 0:
            spans = answers_obj[0].get("spans", [])
            if spans and len(spans) > 0:
                return str(spans[0])

    if "label" in item:
        # hover, exfever 等数据集
        label = item["label"]
        # 将数字标签转换为 Yes/No
        if isinstance(label, int):
            return "Yes" if label == 1 else "No"
        return str(label)

    # 兼容中文数据
    if "claim" in item:
        # 如果有 verdict 字段
        if "verdict" in item:
            return str(item["verdict"])

    return ""


def evaluate_predictions(
    results: List[Dict[str, Any]],
    dataset: str,
    verbose: bool = False
) -> Dict[str, float]:
    """评测预测结果"""

    total = len(results)
    em_scores = []
    f1_scores = []
    yes_no_correct = 0
    has_answer_count = 0

    # 判断是否是事实核查数据集（支持中文）
    # 如果 dataset 为 "auto"，则完全依赖自动检测
    is_fact_check = is_fact_checking_dataset(dataset) if dataset != "auto" else False

    # 自动检测：如果数据包含 claim 字段且答案是 T/F/uncertain，则视为事实核查任务
    if (not is_fact_check or dataset == "auto") and len(results) > 0:
        first_item = results[0]
        if "claim" in first_item or "claim" in first_item.get("original_row", {}):
            ground_truth = get_ground_truth_answer(first_item, dataset)
            if ground_truth and ground_truth.strip().upper() in ['T', 'F', 'UNCERTAIN', '成立', '不成立', '不确定']:
                is_fact_check = True
                if verbose:
                    print("[Info] 检测到事实核查数据集，使用判定结果标准化")

    # 混淆矩阵统计（用于事实核查任务）
    confusion_matrix = {
        'T': {'T': 0, 'F': 0, 'uncertain': 0},
        'F': {'T': 0, 'F': 0, 'uncertain': 0},
        'uncertain': {'T': 0, 'F': 0, 'uncertain': 0}
    }

    for idx, item in enumerate(results):
        # 获取生成的答案
        predicted_answer = item.get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        # 获取标准答案
        ground_truth = get_ground_truth_answer(item, dataset)

        if not ground_truth:
            if verbose:
                print(f"[Warning] 样本 {idx} 缺少标准答案")
            continue

        has_answer_count += 1

        # 计算指标
        em = compute_exact_match(predicted_answer, ground_truth, is_fact_check=is_fact_check)
        f1 = compute_f1(predicted_answer, ground_truth, is_fact_check=is_fact_check)
        em_scores.append(em)
        f1_scores.append(f1)

        # 如果是事实核查任务，记录混淆矩阵
        if is_fact_check:
            pred_normalized = normalize_verdict(predicted_answer)
            truth_normalized = normalize_verdict(ground_truth)
            # 只统计能识别的标签
            if pred_normalized in confusion_matrix and truth_normalized in ['T', 'F', 'uncertain']:
                confusion_matrix[pred_normalized][truth_normalized] += 1

        # 对于分类任务，额外计算准确率
        if dataset in ["hover", "exfever"]:
            acc = compute_yes_no_accuracy(predicted_answer, ground_truth)
            if acc == 1.0:
                yes_no_correct += 1

        if verbose and idx < 10:  # 打印前10个样本的对比
            print(f"\n样本 {idx}:")
            print(f"  问题: {item.get('question') or item.get('claim', '')[:100]}")
            print(f"  预测: {predicted_answer[:100]}")
            print(f"  标准: {ground_truth[:100]}")
            if is_fact_check:
                pred_normalized = normalize_verdict(predicted_answer)
                truth_normalized = normalize_verdict(ground_truth)
                print(f"  标准化: 预测={pred_normalized}, 标准={truth_normalized}")
            print(f"  EM: {em:.2f}, F1: {f1:.2f}")

    # 计算平均指标
    metrics = {}
    if has_answer_count > 0:
        metrics["exact_match"] = sum(em_scores) / len(em_scores) * 100
        metrics["f1"] = sum(f1_scores) / len(f1_scores) * 100
        metrics["evaluated_samples"] = has_answer_count

        if dataset in ["hover", "exfever"]:
            metrics["yes_no_accuracy"] = yes_no_correct / has_answer_count * 100

    metrics["total_samples"] = total

    # 如果是事实核查任务，添加混淆矩阵
    if is_fact_check:
        metrics["confusion_matrix"] = confusion_matrix
        metrics["is_fact_check"] = True

    return metrics


def generate_comparison_excel(
    results: List[Dict[str, Any]],
    metrics: Dict[str, Any],
    output_path: str,
    dataset: str
) -> bool:
    """
    生成包含原始claim、人工评测结果和模型预测结果的对比表格

    Args:
        results: 评测结果列表
        metrics: 评测指标
        output_path: 输出文件路径
        dataset: 数据集类型
    """
    if not EXCEL_AVAILABLE:
        print("错误: 需要安装 pandas 和 openpyxl 来生成Excel文件")
        print("运行: pip install pandas openpyxl")
        return False

    print(f"\n正在生成对比表格: {output_path}")

    # 准备数据
    comparison_data = []
    is_fact_check = metrics.get("is_fact_check", False)

    for i, item in enumerate(results):
        # 提取各个字段
        claim = ""
        if "claim" in item:
            claim = str(item["claim"])
        elif "original_row" in item and "claim" in item["original_row"]:
            claim = str(item["original_row"]["claim"])
        elif "question" in item:
            claim = str(item["question"])
        elif "original_row" in item and "question" in item["original_row"]:
            claim = str(item["original_row"]["question"])

        # 获取人工评测结果
        ground_truth = get_ground_truth_answer(item, dataset)

        # 获取模型预测
        predicted_answer = item.get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        # 标准化判定结果（如果是事实核查任务）
        if is_fact_check:
            ground_truth_normalized = normalize_verdict(ground_truth) if ground_truth else ""
            predicted_normalized = normalize_verdict(predicted_answer) if predicted_answer else ""
            is_match = ground_truth_normalized == predicted_normalized if (ground_truth_normalized and predicted_normalized) else None

            comparison_data.append({
                '序号': i + 1,
                'Claim/问题': claim,
                '人工评测结果': ground_truth,
                '模型预测结果': predicted_answer,
                '标准化-人工': ground_truth_normalized,
                '标准化-模型': predicted_normalized,
                '是否一致': '✓' if is_match else ('✗' if is_match is False else '')
            })
        else:
            # 非事实核查任务，只显示基本信息
            comparison_data.append({
                '序号': i + 1,
                'Question/Claim': claim,
                '标准答案': ground_truth,
                '模型预测': predicted_answer
            })

    # 创建DataFrame
    df = pd.DataFrame(comparison_data)

    # 创建Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='对比结果', index=False)

        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['对比结果']

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        if is_fact_check:
            worksheet.column_dimensions['A'].width = 8   # 序号
            worksheet.column_dimensions['B'].width = 60  # Claim
            worksheet.column_dimensions['C'].width = 15  # 人工结果
            worksheet.column_dimensions['D'].width = 60  # 模型结果
            worksheet.column_dimensions['E'].width = 15  # 标准化-人工
            worksheet.column_dimensions['F'].width = 15  # 标准化-模型
            worksheet.column_dimensions['G'].width = 10  # 是否一致

            # 为不一致的行添加颜色标记
            error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            correct_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')

            for row_idx, row_data in enumerate(comparison_data, start=2):
                if row_data['是否一致'] == '✗':
                    for col in range(1, 8):
                        worksheet.cell(row=row_idx, column=col).fill = error_fill
                elif row_data['是否一致'] == '✓':
                    for col in range(1, 8):
                        worksheet.cell(row=row_idx, column=col).fill = correct_fill
        else:
            worksheet.column_dimensions['A'].width = 8   # 序号
            worksheet.column_dimensions['B'].width = 60  # Question
            worksheet.column_dimensions['C'].width = 40  # 标准答案
            worksheet.column_dimensions['D'].width = 60  # 模型预测

        # 设置所有单元格自动换行
        for row in worksheet.iter_rows(min_row=2, max_row=len(comparison_data) + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    print(f"✓ Excel对比表格已生成: {output_path}")

    # 打印统计信息（如果是事实核查任务）
    if is_fact_check:
        match_count = sum(1 for d in comparison_data if d.get('是否一致') == '✓')
        mismatch_count = sum(1 for d in comparison_data if d.get('是否一致') == '✗')

        print(f"  - 一致样本: {match_count} ({match_count/len(comparison_data)*100:.2f}%)")
        print(f"  - 不一致样本: {mismatch_count} ({mismatch_count/len(comparison_data)*100:.2f}%)")

    return True


def write_to_excel_column(
    results: List[Dict[str, Any]],
    excel_path: str,
    target_column: str = 'R',
    sheet_name: Optional[str] = None
) -> bool:
    """
    将模型预测结果写入现有Excel文件的指定列

    Args:
        results: 评测结果列表
        excel_path: Excel文件路径
        target_column: 目标列（默认R）
        sheet_name: 工作表名称（None表示使用活动工作表）
    """
    if not EXCEL_AVAILABLE:
        print("错误: 需要安装 openpyxl 来操作Excel文件")
        print("运行: pip install openpyxl")
        return False

    from pathlib import Path

    if not Path(excel_path).exists():
        print(f"错误: Excel文件不存在: {excel_path}")
        return False

    print(f"\n正在写入Excel文件: {excel_path}")

    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_path)

    # 选择工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"错误: 工作表 '{sheet_name}' 不存在")
            print(f"可用的工作表: {wb.sheetnames}")
            return False
        ws = wb[sheet_name]
    else:
        ws = wb.active

    print(f"  工作表: {ws.title}")
    print(f"  目标列: {target_column}")

    # 写入表头
    col_letter = target_column.upper()
    ws[f"{col_letter}1"] = "模型预测结果"

    # 写入预测结果
    for i, item in enumerate(results):
        predicted_answer = item.get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        row_number = i + 2  # 从第2行开始（第1行是表头）
        ws[f"{col_letter}{row_number}"] = predicted_answer

    # 保存文件
    output_path = excel_path.replace('.xlsx', '_with_predictions.xlsx')
    wb.save(output_path)

    print(f"✓ 已写入 {len(results)} 条预测结果")
    print(f"✓ 已保存到: {output_path}")

    return True


def main():
    parser = argparse.ArgumentParser(description="评测生成答案与标准答案的匹配度")
    parser.add_argument("--input_file", type=str, required=True,
                        help="生成结果文件路径（包含 final_answer 字段的 JSONL）")
    parser.add_argument("--dataset", type=str, default="auto",
                        choices=["auto", "hotpotqa", "hover", "exfever", "bamboogle", "musique", "2wikimqa"],
                        help="数据集类型（默认：auto 自动检测）")
    parser.add_argument("--verbose", action="store_true",
                        help="是否打印详细对比信息")
    parser.add_argument("--output_file", type=str, default=None,
                        help="保存评测结果的文件路径（可选）")

    # Excel 输出选项
    parser.add_argument("--excel_comparison", type=str, default=None,
                        help="生成Excel对比表格（包含claim、人工结果、模型结果）")
    parser.add_argument("--write_to_excel", type=str, default=None,
                        help="将模型预测结果写入现有Excel文件的指定列")
    parser.add_argument("--excel_column", type=str, default="R",
                        help="写入Excel的目标列（默认：R）")
    parser.add_argument("--excel_sheet", type=str, default=None,
                        help="Excel工作表名称（默认：活动工作表）")

    args = parser.parse_args()

    print("=" * 80)
    print(f"评测配置:")
    print(f"  输入文件: {args.input_file}")
    print(f"  数据集: {args.dataset}")
    print("=" * 80)

    # 加载结果
    print("\n正在加载生成结果...")
    results = load_jsonl(args.input_file)
    print(f"加载了 {len(results)} 条结果\n")

    # 评测
    print("开始评测...")
    metrics = evaluate_predictions(results, args.dataset, args.verbose)

    # 打印结果
    print("\n" + "=" * 80)
    print("评测结果:")
    print("=" * 80)
    print(f"总样本数: {metrics['total_samples']}")
    print(f"有效评测样本数: {metrics.get('evaluated_samples', 0)}")

    if "exact_match" in metrics:
        print(f"\nExact Match (EM): {metrics['exact_match']:.2f}%")
        print(f"F1 Score: {metrics['f1']:.2f}%")

    if "yes_no_accuracy" in metrics:
        print(f"Yes/No Accuracy: {metrics['yes_no_accuracy']:.2f}%")

    # 显示混淆矩阵（如果是事实核查任务）
    if metrics.get("is_fact_check") and "confusion_matrix" in metrics:
        print("\n" + "=" * 80)
        print("混淆矩阵 (Confusion Matrix):")
        print("=" * 80)
        cm = metrics["confusion_matrix"]

        # 表头
        header = "预测 \\ 真实"
        print(f"{header:<15} {'T':>10} {'F':>10} {'uncertain':>10} {'总计':>10}")
        print("-" * 57)

        # 每一行
        for pred_label in ['T', 'F', 'uncertain']:
            row_sum = sum(cm[pred_label].values())
            print(f"{pred_label:<15} {cm[pred_label]['T']:>10} {cm[pred_label]['F']:>10} {cm[pred_label]['uncertain']:>10} {row_sum:>10}")

        # 列总计
        col_t = sum(cm[pred]['T'] for pred in ['T', 'F', 'uncertain'])
        col_f = sum(cm[pred]['F'] for pred in ['T', 'F', 'uncertain'])
        col_u = sum(cm[pred]['uncertain'] for pred in ['T', 'F', 'uncertain'])
        total_count = col_t + col_f + col_u

        print("-" * 57)
        print(f"{'总计':<15} {col_t:>10} {col_f:>10} {col_u:>10} {total_count:>10}")

        # 各类别的 Precision, Recall, F1
        print("\n" + "=" * 80)
        print("各类别指标:")
        print("=" * 80)
        print(f"{'类别':<12} {'Precision':>12} {'Recall':>12} {'F1-Score':>12} {'Support':>12}")
        print("-" * 60)

        for label in ['T', 'F', 'uncertain']:
            # Precision = TP / (TP + FP)
            tp = cm[label][label]
            fp = sum(cm[label][other] for other in ['T', 'F', 'uncertain'] if other != label)
            precision = tp / (tp + fp) * 100 if (tp + fp) > 0 else 0.0

            # Recall = TP / (TP + FN)
            fn = sum(cm[other][label] for other in ['T', 'F', 'uncertain'] if other != label)
            recall = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0.0

            # F1 = 2 * (Precision * Recall) / (Precision + Recall)
            f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0

            # Support = 真实标签为该类的样本数
            support = sum(cm[pred][label] for pred in ['T', 'F', 'uncertain'])

            print(f"{label:<12} {precision:>11.2f}% {recall:>11.2f}% {f1:>11.2f}% {support:>12}")

    print("=" * 80)

    # 保存结果到文件（可选）
    if args.output_file:
        with open(args.output_file, 'w', encoding='utf-8') as f:
            json.dump(metrics, f, ensure_ascii=False, indent=2)
        print(f"\n评测结果已保存到: {args.output_file}")

    # Excel 输出功能
    if args.excel_comparison:
        generate_comparison_excel(
            results=results,
            metrics=metrics,
            output_path=args.excel_comparison,
            dataset=args.dataset
        )

    if args.write_to_excel:
        write_to_excel_column(
            results=results,
            excel_path=args.write_to_excel,
            target_column=args.excel_column,
            sheet_name=args.excel_sheet
        )


if __name__ == "__main__":
    main()
