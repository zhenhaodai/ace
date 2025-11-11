#!/usr/bin/env python3
"""
将模型预测结果写入Excel文件的R列
支持从JSONL结果文件读取模型预测，并更新到指定Excel文件
"""

import argparse
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def load_jsonl(file_path: str):
    """加载JSONL文件"""
    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                data.append(json.loads(line))
    return data


def extract_answer_from_tags(text: str) -> str:
    """从 <answer></answer> 标签中提取答案"""
    import re
    match = re.search(r'<answer>\s*(.*?)\s*</answer>', text, re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()
    return text.strip()


def write_predictions_to_excel(
    excel_path: str,
    predictions_path: str,
    target_column: str = 'R',
    sheet_name: str = None,
    header_row: int = 1,
    start_data_row: int = 2
):
    """
    将预测结果写入Excel文件的指定列

    Args:
        excel_path: Excel文件路径
        predictions_path: 包含预测结果的JSONL文件路径
        target_column: 目标列（默认R列）
        sheet_name: 工作表名称（None表示使用活动工作表）
        header_row: 表头所在行号（默认第1行）
        start_data_row: 数据开始行号（默认第2行）
    """

    # 检查文件是否存在
    if not Path(excel_path).exists():
        print(f"错误: Excel文件不存在: {excel_path}")
        return False

    if not Path(predictions_path).exists():
        print(f"错误: 预测结果文件不存在: {predictions_path}")
        return False

    # 加载预测结果
    print(f"正在加载预测结果: {predictions_path}")
    predictions = load_jsonl(predictions_path)
    print(f"加载了 {len(predictions)} 条预测结果")

    # 打开Excel文件
    print(f"正在打开Excel文件: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    # 选择工作表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"错误: 工作表 '{sheet_name}' 不存在")
            print(f"可用的工作表: {wb.sheetnames}")
            return False
        ws = wb[sheet_name]
    else:
        ws = wb.active

    print(f"使用工作表: {ws.title}")

    # 写入表头
    col_letter = target_column.upper()
    header_cell = f"{col_letter}{header_row}"
    ws[header_cell] = "模型预测结果"
    print(f"已写入表头到 {header_cell}")

    # 写入预测结果
    write_count = 0
    for i, pred_item in enumerate(predictions):
        # 提取预测答案
        predicted_answer = pred_item.get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        # 计算行号
        row_number = start_data_row + i

        # 写入单元格
        cell_address = f"{col_letter}{row_number}"
        ws[cell_address] = predicted_answer
        write_count += 1

        if (i + 1) % 100 == 0:
            print(f"已写入 {i + 1} 条结果...")

    print(f"共写入 {write_count} 条预测结果到列 {col_letter}")

    # 保存文件
    output_path = excel_path.replace('.xlsx', '_with_predictions.xlsx')
    print(f"正在保存到: {output_path}")
    wb.save(output_path)
    print(f"✓ 成功保存！")

    return True


def main():
    parser = argparse.ArgumentParser(
        description="将模型预测结果写入Excel文件的指定列",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  # 写入R列（默认）
  python write_predictions_to_excel.py --excel data.xlsx --predictions results.jsonl

  # 写入S列
  python write_predictions_to_excel.py --excel data.xlsx --predictions results.jsonl --column S

  # 指定工作表
  python write_predictions_to_excel.py --excel data.xlsx --predictions results.jsonl --sheet "Sheet1"
        """
    )

    parser.add_argument("--excel", type=str, required=True,
                        help="Excel文件路径")
    parser.add_argument("--predictions", type=str, required=True,
                        help="预测结果JSONL文件路径")
    parser.add_argument("--column", type=str, default="R",
                        help="目标列（默认: R）")
    parser.add_argument("--sheet", type=str, default=None,
                        help="工作表名称（默认: 活动工作表）")
    parser.add_argument("--header-row", type=int, default=1,
                        help="表头所在行号（默认: 1）")
    parser.add_argument("--start-row", type=int, default=2,
                        help="数据开始行号（默认: 2）")

    args = parser.parse_args()

    print("=" * 80)
    print("将模型预测结果写入Excel")
    print("=" * 80)
    print(f"Excel文件: {args.excel}")
    print(f"预测结果: {args.predictions}")
    print(f"目标列: {args.column}")
    print(f"工作表: {args.sheet or '(活动工作表)'}")
    print("=" * 80)
    print()

    success = write_predictions_to_excel(
        excel_path=args.excel,
        predictions_path=args.predictions,
        target_column=args.column,
        sheet_name=args.sheet,
        header_row=args.header_row,
        start_data_row=args.start_row
    )

    if success:
        print("\n✓ 完成！")
        sys.exit(0)
    else:
        print("\n✗ 失败！")
        sys.exit(1)


if __name__ == "__main__":
    main()
