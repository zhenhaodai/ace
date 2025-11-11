#!/usr/bin/env python3
"""
生成包含原始claim、人工评测结果和模型预测结果的对比表格
支持从JSONL结果文件和Excel文件生成新的对比表格
"""

import argparse
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd


def load_jsonl(file_path: str):
    """加载JSONL文件"""
    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                data.append(json.loads(line))
    return data


def extract_answer_from_tags(text: str) -> str:
    """从 <answer></answer> 标签中提取答案"""
    import re
    match = re.search(r'<answer>\s*(.*?)\s*</answer>', text, re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()
    return text.strip()


def normalize_verdict(s: str) -> str:
    """标准化判定结果（支持中英文）"""
    s = str(s).strip().lower()

    # True/成立 相关
    if s in ['t', 'true', '成立', '正确', '支持', 'yes']:
        return 'T'

    # False/不成立 相关
    if s in ['f', 'false', '不成立', '错误', '不支持', 'no']:
        return 'F'

    # Uncertain/不确定 相关
    if s in ['uncertain', 'u', '不确定', '证据不足', '无法判断', '未知']:
        return 'uncertain'

    return s


def get_ground_truth_from_item(item: dict) -> str:
    """从JSONL项中提取人工评测结果"""
    # 检查 original_row
    if "original_row" in item:
        original = item["original_row"]
        for field in ["人工评测结果", "标准答案", "答案", "label", "answer"]:
            if field in original and original[field]:
                answer = original[field]
                if answer != answer:  # NaN check
                    continue
                return str(answer)

    # 检查根节点
    for field in ["人工评测结果", "标准答案", "答案", "answer", "label"]:
        if field in item and item[field]:
            answer = item[field]
            if answer != answer:  # NaN check
                continue
            return str(answer)

    return ""


def get_claim_from_item(item: dict) -> str:
    """从JSONL项中提取claim"""
    # 检查 original_row
    if "original_row" in item:
        original = item["original_row"]
        if "claim" in original:
            return str(original["claim"])

    # 检查根节点
    if "claim" in item:
        return str(item["claim"])

    # 如果没有claim，尝试question
    if "question" in item:
        return str(item["question"])

    if "original_row" in item and "question" in item["original_row"]:
        return str(item["original_row"]["question"])

    return ""


def generate_comparison_table_from_jsonl(
    predictions_path: str,
    output_path: str,
    format: str = 'xlsx'
):
    """
    从JSONL文件生成对比表格

    Args:
        predictions_path: 包含预测结果的JSONL文件路径
        output_path: 输出文件路径
        format: 输出格式（xlsx或csv）
    """

    # 加载预测结果
    print(f"正在加载预测结果: {predictions_path}")
    predictions = load_jsonl(predictions_path)
    print(f"加载了 {len(predictions)} 条预测结果")

    # 准备数据
    comparison_data = []
    for i, item in enumerate(predictions):
        # 提取各个字段
        claim = get_claim_from_item(item)
        ground_truth = get_ground_truth_from_item(item)
        predicted_answer = item.get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        # 标准化判定结果
        ground_truth_normalized = normalize_verdict(ground_truth) if ground_truth else ""
        predicted_normalized = normalize_verdict(predicted_answer) if predicted_answer else ""

        # 判断是否一致
        is_match = ground_truth_normalized == predicted_normalized if (ground_truth_normalized and predicted_normalized) else None

        comparison_data.append({
            '序号': i + 1,
            'Claim/问题': claim,
            '人工评测结果': ground_truth,
            '模型预测结果': predicted_answer,
            '标准化-人工': ground_truth_normalized,
            '标准化-模型': predicted_normalized,
            '是否一致': '✓' if is_match else ('✗' if is_match is False else '')
        })

    # 创建DataFrame
    df = pd.DataFrame(comparison_data)

    # 保存文件
    if format.lower() == 'xlsx':
        print(f"正在生成Excel文件: {output_path}")

        # 创建Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='对比结果', index=False)

            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['对比结果']

            # 设置表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            worksheet.column_dimensions['A'].width = 8   # 序号
            worksheet.column_dimensions['B'].width = 60  # Claim
            worksheet.column_dimensions['C'].width = 15  # 人工结果
            worksheet.column_dimensions['D'].width = 60  # 模型结果
            worksheet.column_dimensions['E'].width = 15  # 标准化-人工
            worksheet.column_dimensions['F'].width = 15  # 标准化-模型
            worksheet.column_dimensions['G'].width = 10  # 是否一致

            # 为不一致的行添加颜色标记
            error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            correct_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')

            for row_idx, row_data in enumerate(comparison_data, start=2):
                if row_data['是否一致'] == '✗':
                    for col in range(1, 8):
                        worksheet.cell(row=row_idx, column=col).fill = error_fill
                elif row_data['是否一致'] == '✓':
                    for col in range(1, 8):
                        worksheet.cell(row=row_idx, column=col).fill = correct_fill

            # 设置所有单元格自动换行
            for row in worksheet.iter_rows(min_row=2, max_row=len(comparison_data) + 1):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        print(f"✓ Excel文件生成成功！")

    elif format.lower() == 'csv':
        print(f"正在生成CSV文件: {output_path}")
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"✓ CSV文件生成成功！")

    # 打印统计信息
    match_count = sum(1 for d in comparison_data if d['是否一致'] == '✓')
    mismatch_count = sum(1 for d in comparison_data if d['是否一致'] == '✗')

    print("\n" + "=" * 80)
    print("统计信息:")
    print("=" * 80)
    print(f"总样本数: {len(comparison_data)}")
    print(f"一致数量: {match_count} ({match_count/len(comparison_data)*100:.2f}%)")
    print(f"不一致数量: {mismatch_count} ({mismatch_count/len(comparison_data)*100:.2f}%)")
    print("=" * 80)

    return True


def generate_comparison_table_from_excel(
    excel_path: str,
    predictions_path: str,
    output_path: str,
    human_column: str = None,
    claim_column: str = None,
    sheet_name: str = None
):
    """
    从Excel文件和JSONL预测结果生成对比表格

    Args:
        excel_path: 原始Excel文件路径（包含人工评测结果）
        predictions_path: JSONL预测结果文件路径
        output_path: 输出文件路径
        human_column: 人工评测结果所在列名
        claim_column: Claim所在列名
        sheet_name: 工作表名称
    """

    # 加载Excel
    print(f"正在加载Excel文件: {excel_path}")
    if sheet_name:
        df_excel = pd.read_excel(excel_path, sheet_name=sheet_name)
    else:
        df_excel = pd.read_excel(excel_path)

    print(f"Excel包含 {len(df_excel)} 行数据")
    print(f"列名: {list(df_excel.columns)}")

    # 自动检测列名（如果未指定）
    if not human_column:
        for col in df_excel.columns:
            if '人工' in str(col) or '评测' in str(col) or 'label' in str(col).lower():
                human_column = col
                break

    if not claim_column:
        for col in df_excel.columns:
            if 'claim' in str(col).lower() or '问题' in str(col):
                claim_column = col
                break

    if not human_column or not claim_column:
        print(f"警告: 未能自动检测列名")
        print(f"  人工评测结果列: {human_column}")
        print(f"  Claim列: {claim_column}")

    # 加载预测结果
    print(f"正在加载预测结果: {predictions_path}")
    predictions = load_jsonl(predictions_path)
    print(f"加载了 {len(predictions)} 条预测结果")

    # 准备数据
    comparison_data = []
    for i in range(min(len(df_excel), len(predictions))):
        # 从Excel提取
        claim = df_excel.iloc[i][claim_column] if claim_column else ""
        ground_truth = df_excel.iloc[i][human_column] if human_column else ""

        # 从JSONL提取预测
        predicted_answer = predictions[i].get("final_answer", "")
        predicted_answer = extract_answer_from_tags(predicted_answer)

        # 标准化
        ground_truth_normalized = normalize_verdict(str(ground_truth)) if pd.notna(ground_truth) else ""
        predicted_normalized = normalize_verdict(predicted_answer) if predicted_answer else ""

        # 判断是否一致
        is_match = ground_truth_normalized == predicted_normalized if (ground_truth_normalized and predicted_normalized) else None

        comparison_data.append({
            '序号': i + 1,
            'Claim/问题': claim,
            '人工评测结果': ground_truth,
            '模型预测结果': predicted_answer,
            '标准化-人工': ground_truth_normalized,
            '标准化-模型': predicted_normalized,
            '是否一致': '✓' if is_match else ('✗' if is_match is False else '')
        })

    # 生成Excel
    df = pd.DataFrame(comparison_data)

    print(f"正在生成对比表格: {output_path}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='对比结果', index=False)

        workbook = writer.book
        worksheet = writer.sheets['对比结果']

        # 样式设置（同上）
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 60
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 60
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 10

        error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        correct_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')

        for row_idx, row_data in enumerate(comparison_data, start=2):
            if row_data['是否一致'] == '✗':
                for col in range(1, 8):
                    worksheet.cell(row=row_idx, column=col).fill = error_fill
            elif row_data['是否一致'] == '✓':
                for col in range(1, 8):
                    worksheet.cell(row=row_idx, column=col).fill = correct_fill

        for row in worksheet.iter_rows(min_row=2, max_row=len(comparison_data) + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    print(f"✓ 对比表格生成成功！")

    # 打印统计
    match_count = sum(1 for d in comparison_data if d['是否一致'] == '✓')
    mismatch_count = sum(1 for d in comparison_data if d['是否一致'] == '✗')

    print("\n" + "=" * 80)
    print("统计信息:")
    print("=" * 80)
    print(f"总样本数: {len(comparison_data)}")
    print(f"一致数量: {match_count} ({match_count/len(comparison_data)*100:.2f}%)")
    print(f"不一致数量: {mismatch_count} ({mismatch_count/len(comparison_data)*100:.2f}%)")
    print("=" * 80)

    return True


def main():
    parser = argparse.ArgumentParser(
        description="生成包含原始claim、人工评测结果和模型预测结果的对比表格",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:

  # 从JSONL生成对比表格
  python generate_comparison_table.py --predictions results.jsonl --output comparison.xlsx

  # 从Excel和JSONL生成对比表格
  python generate_comparison_table.py --excel data.xlsx --predictions results.jsonl --output comparison.xlsx

  # 指定列名
  python generate_comparison_table.py --excel data.xlsx --predictions results.jsonl \\
      --human-column "人工评测结果" --claim-column "claim" --output comparison.xlsx

  # 生成CSV格式
  python generate_comparison_table.py --predictions results.jsonl --output comparison.csv --format csv
        """
    )

    parser.add_argument("--predictions", type=str, required=True,
                        help="预测结果JSONL文件路径")
    parser.add_argument("--output", type=str, required=True,
                        help="输出文件路径")
    parser.add_argument("--excel", type=str, default=None,
                        help="原始Excel文件路径（可选，如果提供则从Excel读取人工评测结果）")
    parser.add_argument("--format", type=str, default="xlsx", choices=["xlsx", "csv"],
                        help="输出格式（默认: xlsx）")
    parser.add_argument("--sheet", type=str, default=None,
                        help="Excel工作表名称（默认: 第一个工作表）")
    parser.add_argument("--human-column", type=str, default=None,
                        help="人工评测结果所在列名（自动检测）")
    parser.add_argument("--claim-column", type=str, default=None,
                        help="Claim所在列名（自动检测）")

    args = parser.parse_args()

    print("=" * 80)
    print("生成对比表格")
    print("=" * 80)
    print(f"预测结果: {args.predictions}")
    print(f"输出文件: {args.output}")
    if args.excel:
        print(f"Excel文件: {args.excel}")
    print(f"输出格式: {args.format}")
    print("=" * 80)
    print()

    try:
        if args.excel:
            # 从Excel和JSONL生成
            success = generate_comparison_table_from_excel(
                excel_path=args.excel,
                predictions_path=args.predictions,
                output_path=args.output,
                human_column=args.human_column,
                claim_column=args.claim_column,
                sheet_name=args.sheet
            )
        else:
            # 仅从JSONL生成
            success = generate_comparison_table_from_jsonl(
                predictions_path=args.predictions,
                output_path=args.output,
                format=args.format
            )

        if success:
            print("\n✓ 完成！")
            sys.exit(0)
        else:
            print("\n✗ 失败！")
            sys.exit(1)

    except Exception as e:
        print(f"\n✗ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
